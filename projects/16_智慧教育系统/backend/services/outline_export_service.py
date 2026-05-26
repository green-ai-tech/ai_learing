"""教学大纲 XLSX 导出服务。"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from backend.models.outline import OutlineModel


STORAGE_ROOT = Path(__file__).resolve().parents[1] / "storage"
XLSX_DIR = STORAGE_ROOT / "xlsx"


class OutlineExportService:
    """将结构化大纲写入 XLSX。"""

    def generate_xlsx(self, outline: OutlineModel) -> Path:
        if not outline.outline_json:
            raise ValueError("大纲内容为空，无法导出 XLSX")

        XLSX_DIR.mkdir(parents=True, exist_ok=True)
        file_path = XLSX_DIR / f"outline_{outline.id}_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"

        wb = Workbook()
        wb.remove(wb.active)

        self._write_basic_sheet(wb, outline.outline_json)
        self._write_list_sheet(wb, "教学目标", ["序号", "目标"], outline.outline_json.get("teaching_goals", []))
        self._write_chapter_sheet(wb, outline.outline_json)
        self._write_list_sheet(
            wb,
            "教学要求",
            ["序号", "要求"],
            outline.outline_json.get("teaching_requirements", []),
        )
        self._write_list_sheet(
            wb,
            "考核方式",
            ["序号", "方式"],
            outline.outline_json.get("assessment_methods", []),
        )

        wb.save(file_path)
        return file_path

    def _write_basic_sheet(self, wb: Workbook, data: dict[str, Any]) -> None:
        ws = wb.create_sheet("基本信息")
        rows = [
            ("课程名称", data.get("course_title", "")),
            ("课程介绍", data.get("course_description", "")),
            ("目标学生", data.get("target_students", "")),
            ("学段", data.get("stage", "")),
            ("总课时", data.get("total_hours", "")),
            ("难度", data.get("difficulty", "")),
            ("教学重点", "\n".join(data.get("key_points", []))),
            ("教学难点", "\n".join(data.get("difficult_points", []))),
            ("参考资料", "\n".join(data.get("references", []))),
        ]
        ws.append(["字段", "内容"])
        for row in rows:
            ws.append(row)
        self._style_sheet(ws)
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 80

    def _write_list_sheet(
        self,
        wb: Workbook,
        title: str,
        headers: list[str],
        items: list[str],
    ) -> None:
        ws = wb.create_sheet(title)
        ws.append(headers)
        for index, item in enumerate(items, 1):
            ws.append([index, item])
        self._style_sheet(ws)
        ws.column_dimensions["A"].width = 10
        ws.column_dimensions["B"].width = 90

    def _write_chapter_sheet(self, wb: Workbook, data: dict[str, Any]) -> None:
        ws = wb.create_sheet("章节安排")
        ws.append(["章", "节", "课时", "描述", "知识点", "教学方法", "评价方式"])
        for chapter in data.get("chapters", []):
            sections = chapter.get("sections") or []
            if not sections:
                ws.append([
                    chapter.get("title", ""),
                    "",
                    chapter.get("hours", ""),
                    chapter.get("description", ""),
                    "",
                    "",
                    "",
                ])
                continue
            for section in sections:
                knowledge_points = [
                    point.get("name", "")
                    for point in section.get("knowledge_points", [])
                    if point.get("name")
                ]
                ws.append(
                    [
                        chapter.get("title", ""),
                        section.get("title", ""),
                        section.get("hours", ""),
                        section.get("description", ""),
                        "\n".join(knowledge_points),
                        "\n".join(section.get("teaching_methods", [])),
                        section.get("assessment", ""),
                    ]
                )
        self._style_sheet(ws)
        widths = [24, 28, 10, 48, 38, 26, 26]
        for index, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(index)].width = width

    def _style_sheet(self, ws) -> None:
        header_fill = PatternFill("solid", fgColor="1F2937")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

